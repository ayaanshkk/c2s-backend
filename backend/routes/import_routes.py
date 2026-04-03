"""
Bulk Import Route for Energy Customers
Handles Excel/CSV uploads and bulk insertion into database
"""

from flask import Blueprint, request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime
from sqlalchemy import and_, or_, text
from flask_jwt_extended import jwt_required, get_jwt_identity
import logging
import tempfile

from ..models import (
    Client_Master, Project_Details, Energy_Contract_Master,
    Supplier_Master, Employee_Master, Services_Master
)
from .auth_helpers import token_required
from ..db import SessionLocal
from .leads_import_handler import import_leads_handler, download_leads_template_handler


logger = logging.getLogger(__name__)
import_bp = Blueprint('import', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
UPLOAD_FOLDER = '/tmp/uploads'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_tenant_id_from_user(user):
    """Get tenant_id from authenticated user - match customer_routes (JWT tenant_id first)"""
    if hasattr(user, 'tenant_id') and user.tenant_id is not None:
        return user.tenant_id
    session = SessionLocal()
    try:
        employee = session.query(Employee_Master).filter_by(employee_id=user.employee_id).first()
        return employee.tenant_id if employee else None
    finally:
        session.close()

def find_supplier_id(supplier_name, session):
    """Find supplier ID by name (case-insensitive, fuzzy matching)"""
    if not supplier_name or pd.isna(supplier_name):
        return None
    
    supplier_name = str(supplier_name).strip()
    
    # Try exact match first
    supplier = session.query(Supplier_Master).filter(
        Supplier_Master.supplier_company_name.ilike(f'%{supplier_name}%')
    ).first()
    
    if supplier:
        return supplier.supplier_id
    
    # Try extracting name before parenthesis (e.g., "THRE (Corona Energy)" -> "THRE")
    if '(' in supplier_name:
        short_name = supplier_name.split('(')[0].strip()
        supplier = session.query(Supplier_Master).filter(
            Supplier_Master.supplier_company_name.ilike(f'%{short_name}%')
        ).first()
        if supplier:
            return supplier.supplier_id
    
    # Try extracting name in parenthesis (e.g., "THRE (Corona Energy)" -> "Corona Energy")
    if '(' in supplier_name and ')' in supplier_name:
        paren_name = supplier_name.split('(')[1].split(')')[0].strip()
        supplier = session.query(Supplier_Master).filter(
            Supplier_Master.supplier_company_name.ilike(f'%{paren_name}%')
        ).first()
        if supplier:
            return supplier.supplier_id
    
    # Try first word only (e.g., "British Gas Business" -> "British")
    first_word = supplier_name.split()[0] if supplier_name.split() else supplier_name
    if len(first_word) > 3:  # Only try if word is longer than 3 chars
        supplier = session.query(Supplier_Master).filter(
            Supplier_Master.supplier_company_name.ilike(f'{first_word}%')
        ).first()
        if supplier:
            return supplier.supplier_id
    
    return None


def get_or_create_supplier(supplier_name, session):
    """Get existing supplier or create new one if doesn't exist"""
    if not supplier_name or pd.isna(supplier_name):
        return 1  # Return default supplier_id
    
    supplier_name = str(supplier_name).strip()
    
    # Try to find existing supplier
    supplier_id = find_supplier_id(supplier_name, session)
    if supplier_id:
        return supplier_id
    
    # Supplier doesn't exist - create it
    try:
        new_supplier = Supplier_Master(
            supplier_company_name=supplier_name,
            supplier_contact_name='Auto-imported',
            supplier_provisions=3,  # Default: Electricity & Gas
            created_at=datetime.utcnow()
        )
        session.add(new_supplier)
        session.flush()
        
        current_app.logger.info(f"✨ Created new supplier: {supplier_name} (ID: {new_supplier.supplier_id})")
        return new_supplier.supplier_id
    except Exception as e:
        current_app.logger.error(f"Failed to create supplier {supplier_name}: {e}")
        return 1  # Fallback to default


def get_or_create_service(tenant_id, session):
    """Get existing default service or create one if doesn't exist"""
    # Try to find existing service for this tenant
    service = session.query(Services_Master).filter_by(
        tenant_id=tenant_id,
        service_title='Default Energy Service'
    ).first()
    
    if service:
        return service.service_id
    
    # Try to get any service for this tenant
    service = session.query(Services_Master).filter_by(tenant_id=tenant_id).first()
    if service:
        return service.service_id
    
    # No service exists - create default one
    try:
        new_service = Services_Master(
            tenant_id=tenant_id,
            service_title='Default Energy Service',
            service_description='Auto-created default service for energy contracts',
            service_rate=0.0,
            currency_id=1,
            supplier_id=None,
            date_from=None,
            date_to=None,
            created_at=datetime.utcnow(),
            service_code='DEFAULT'
        )
        session.add(new_service)
        session.flush()
        
        current_app.logger.info(f"✨ Created default service for tenant {tenant_id} (ID: {new_service.service_id})")
        return new_service.service_id
    except Exception as e:
        current_app.logger.error(f"Failed to create default service: {e}")
        return 1  # Fallback to ID 1

# def parse_date(date_value):
#     """Parse date from various formats - prioritize DD/MM/YYYY (UK format)"""
#     if pd.isna(date_value) or not date_value:
#         return None
    
#     if isinstance(date_value, datetime):
#         return date_value.date()
    
#     date_str = str(date_value).strip()
    
#     # ✅ UPDATED: Prioritize UK date formats (DD/MM/YYYY)
#     date_formats = [
#         '%d/%m/%Y',      
#         '%d-%m-%Y',      
#         '%d.%m.%Y',      
#         '%d %b %Y',      
#         '%d %B %Y',      
#         '%Y-%m-%d',      
#         '%m/%d/%Y',      
#         '%Y/%m/%d',
#     ]
    
#     for fmt in date_formats:
#         try:
#             return datetime.strptime(date_str, fmt).date()
#         except ValueError:
#             continue
    
#     return None

def parse_date(date_value):
    """Parse date from various formats - prioritize DD/MM/YYYY (UK format)"""
    if pd.isna(date_value) or not date_value or date_value == '':
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    date_str = str(date_value).strip()
    
    # Skip empty or 'nan' strings
    if not date_str or date_str.lower() == 'nan':
        return None
    
    # ✅ Prioritize UK date formats (DD/MM/YYYY) + datetime formats
    date_formats = [
        '%Y-%m-%d %H:%M:%S',  # ✅ ADD THIS FIRST - Excel datetime format
        '%d/%m/%Y',      
        '%d-%m-%Y',      
        '%d.%m.%Y',      
        '%d %b %Y',      
        '%d %B %Y',      
        '%Y-%m-%d',      
        '%m/%d/%Y',      
        '%Y/%m/%d',
    ]
    
    for fmt in date_formats:
        try:
            parsed = datetime.strptime(date_str, fmt).date()
            return parsed
        except ValueError:
            continue
    
    return None

# def parse_number(value):
#     """Parse number from string (handles commas, etc.)"""
#     if pd.isna(value) or not value:
#         return None
    
#     try:
#         # Remove commas and convert to float
#         cleaned = str(value).replace(',', '').strip()
#         return float(cleaned) if cleaned else None
#     except (ValueError, AttributeError):
#         return None

def parse_number(value):
    """Parse number from string (handles commas, etc.)"""
    if pd.isna(value) or not value or value == '':
        return None
    
    try:
        # Remove commas and convert to float
        cleaned = str(value).replace(',', '').strip()
        
        # ✅ Handle empty string after cleaning
        if not cleaned or cleaned == 'nan':
            return None
            
        return float(cleaned) if cleaned else None
    except (ValueError, AttributeError):
        return None

def safe_str(value):
    """Convert value to clean string, remove .0 suffix from numeric strings"""
    if pd.isna(value) or value is None or value == '':
        return ''
    str_value = str(value).strip()
    if str_value.endswith('.0') and str_value[:-2].replace('.', '', 1).isdigit():
        str_value = str_value[:-2]
    return str_value


@import_bp.route('/energy-customers', methods=['POST', 'OPTIONS'])
@token_required
def import_energy_customers():
    """
    Bulk import energy customers from Excel/CSV file with optional assignment
    ⚡ HANDLES UNLIMITED RECORDS: Individual commits prevent timeouts
    """
    print("\n\n🔥🔥🔥 IMPORT FUNCTION CALLED! 🔥🔥🔥\n\n")
    
    if request.method == 'OPTIONS':
        print("OPTIONS request - returning")
        return jsonify({}), 200
    
    print("Creating session...")
    session = SessionLocal()
    
    print("Setting up logging...")
    import logging
    sql_logger = logging.getLogger('sqlalchemy.engine')
    original_level = sql_logger.level
    sql_logger.setLevel(logging.WARNING)
    
    try:
        print("Checking for file...")
        if 'file' not in request.files:
            print("ERROR: No file uploaded")
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv'}), 400
        
        # Get tenant and user info
        tenant_id = get_tenant_id_from_user(request.current_user)
        if not tenant_id:
            return jsonify({'error': 'Tenant not found for user'}), 400
        
        employee_id = request.current_user.employee_id

        # GET ASSIGNED EMPLOYEE ID FROM FORM DATA
        assigned_employee_id = request.form.get('assigned_employee_id', type=int)
        opportunity_owner_id = assigned_employee_id if assigned_employee_id else employee_id
        
        # Get employee name for success message
        assigned_employee_name = None
        if assigned_employee_id:
            assigned_employee = session.query(Employee_Master).filter_by(
                employee_id=assigned_employee_id,
                tenant_id=tenant_id
            ).first()
            if assigned_employee:
                assigned_employee_name = assigned_employee.employee_name
            else:
                return jsonify({'error': f'Invalid employee ID: {assigned_employee_id}'}), 400

        print(f"\n{'='*60}")
        print(f"📥 BULK IMPORT STARTED")
        print(f"{'='*60}")
        print(f"   Tenant ID: {tenant_id}")
        print(f"   Uploaded by: Employee ID {employee_id}")
        print(f"   Assigned to: {assigned_employee_name or 'Uploader'} (ID: {opportunity_owner_id})")
        print(f"{'='*60}\n")

        # Service filter
        service_param = request.args.get('service', 'utilities')
        service_id_map = {
            'utilities': 1,
            'electricity': 1, 
            'water': 2,
            'gas': 3
        }
        import_service_id = service_id_map.get(service_param.strip().lower(), 1)
        
        # Read file
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower()

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name
            
            if file_ext == 'csv':
                df = pd.read_csv(tmp_path, encoding='utf-8-sig', dtype=str)
            else:
                try:
                    df = pd.read_excel(tmp_path, engine='openpyxl', dtype=str)
                except Exception:
                    df = pd.read_excel(tmp_path, engine='xlrd', dtype=str)
            
            os.unlink(tmp_path)
            
        except Exception as e:
            print(f"❌ Failed to read file: {str(e)}")
            return jsonify({'error': f'Failed to read file: {str(e)}'}), 400
                    
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace('_', ' ').str.replace(r'\s+', ' ', regex=True)
        
        # Column mapping
        column_map = {
            # Contact Information
            'client_name': ['client name', 'business name', 'company name'],
            'trading_name': ['trading name', 'business', 'company'],
            'main_contact': ['main contact', 'contact person', 'contact'],
            'position': ['position', 'role', 'title'],
            'tel_no': ['tel no', 'phone', 'telephone', 'tel'],
            'mobile_no': ['mobile no', 'mobile', 'cell'],
            'email': ['email', 'e-mail'],
            
            # Site/Property Information
            'site_name': ['site name', 'site'],
            'month_sold': ['month sold', 'sale month'],
            'house_name': ['house name'],
            'house_number': ['house number', 'house no'],
            'door_number': ['door number'],
            
            # Address Fields
            'address_line_1': ['address line 1', 'address 1', 'street'],
            'address_line_2': ['address line 2', 'address 2'],
            'address_line_3': ['address line 3', 'address 3'],
            'town': ['town', 'city'],
            'county': ['county', 'region'],
            
            # ✅ CORRECTED: Both postcode fields map to same column
            'postcode': ['postcode', 'post code', 'zip', 'home post code', 'home postcode'],
            
            # Home Address (separate from trading address)
            'home_door_number': ['home door number', 'home door no'],
            'home_street': ['home street'],
            # ❌ REMOVED: 'home_post_code' - using postcode instead
            
            # MPAN/Meter Information
            'mpan_top': ['mpan top', 'mpan core'],
            'mpan_bottom': ['mpan bottom', 'mpan llf'],
            'data_source': ['data source'],
            
            # Supplier Information
            'old_supplier': ['old supplier'],
            'supplier': ['supplier', 'supplier name'],
            
            # Contract Details
            'payment_type': ['payment type'],
            'net_notch': ['net notch'],
            'term_sold': ['term sold', 'in contract', 'contract length'],
            'agent_sold': ['agent sold'],
            'start_date': ['start date', 'contract start'],
            'contract_end': ['contract end', 'end date', 'expiry'],
            
            # Charges
            'stand_charge': ['stand charge', 'standing charge'],
            'rate_1': ['rate 1', 'unit rate', 'rate'],
            'rate_2': ['rate 2'],
            'rate_3': ['rate 3'],
            'aggregator': ['aggregator'],
            'annual_usage': ['annual usage', 'usage', 'kwh'],
            'comms_paid': ['comms paid', 'commission'],
            
            # Company/Trading Information
            'trading_type': ['trading type'],
            'company_number': ['company number', 'co number'],
            'date_of_birth': ['date of birth', 'dob'],
            'charity_ltd_company_number': ['charity/ltd company number', 'charity number'],
            
            # Banking Information
            'bank_name': ['bank name', 'bank'],
            'ac_number': ['ac number', 'account number'],
            'sort_code': ['sort code'],
            
            # Partner Information
            'partner_details': ['partner details', 'partner'],
            'partner_dob': ['partner date of birth', 'partner dob'],
            
            # Customer Verification
            'credit_score': ['credit score'],
            'password': ['password'],
        }
        
        actual_columns = {}
        for field, possible_names in column_map.items():
            for col in df.columns:
                if col in possible_names:
                    actual_columns[field] = col
                    break
        
        # PRE-LOAD SUPPLIERS
        suppliers_dict = {}
        suppliers = session.query(Supplier_Master).all()
        for s in suppliers:
            suppliers_dict[s.supplier_company_name.lower().strip()] = s.supplier_id
        
        print(f"📊 Loaded {len(suppliers_dict)} suppliers for matching")
        
        # PRE-LOAD EXISTING MPANs
        existing_mpans = {}  
        existing_contracts_query = session.query(
            Energy_Contract_Master,
            Client_Master.tenant_id,
            Project_Details.assigned_employee_id,
            Employee_Master.employee_name,
            Client_Master.client_company_name,
            Client_Master.is_archived
        ).join(
            Project_Details, Energy_Contract_Master.project_id == Project_Details.project_id
        ).join(
            Client_Master, Project_Details.client_id == Client_Master.client_id
        ).outerjoin(
            Employee_Master, Project_Details.assigned_employee_id == Employee_Master.employee_id  
        ).all()

        for contract, contract_tenant_id, assigned_emp_id, emp_name, company_name, is_archived in existing_contracts_query:
            if contract.mpan_number:
                mpan_key = contract.mpan_number.strip().lower()
                
                # ✅ CRITICAL FIX: Store as LIST to handle multiple tenants with same MPAN
                if mpan_key not in existing_mpans:
                    existing_mpans[mpan_key] = []
                
                existing_mpans[mpan_key].append({
                    'contract': contract,
                    'tenant_id': contract_tenant_id,
                    'assigned_to_id': assigned_emp_id,
                    'assigned_to_name': emp_name or 'Unassigned',
                    'company_name': company_name,
                    'is_archived': is_archived
                })

        print(f"📊 Loaded {sum(len(v) for v in existing_mpans.values())} total contracts across {len(existing_mpans)} unique MPANs for cross-tenant duplicate checking")

        # Track duplicate info for final report
        duplicate_details = []
        cross_tenant_duplicates = []
        
        # PROCESS EACH RECORD
        total_rows = len(df)
        success_count = 0
        error_count = 0
        duplicate_count = 0
        errors = []
        BATCH_SIZE = 50
        
        print(f"📊 Starting import of {total_rows} rows (individual commits)")
        
        for index, row in df.iterrows():
            try:
                # Extract data
                client_name = safe_str(row.get(actual_columns.get('client_name', ''), ''))
                trading_name = safe_str(row.get(actual_columns.get('trading_name', ''), ''))
                main_contact = safe_str(row.get(actual_columns.get('main_contact', ''), ''))
                position = safe_str(row.get(actual_columns.get('position', ''), ''))
                tel_no = safe_str(row.get(actual_columns.get('tel_no', ''), ''))
                mobile_no = safe_str(row.get(actual_columns.get('mobile_no', ''), ''))
                email = safe_str(row.get(actual_columns.get('email', ''), ''))
                site_name = safe_str(row.get(actual_columns.get('site_name', ''), ''))

                # Address fields
                address_line_1 = safe_str(row.get(actual_columns.get('address_line_1', ''), ''))
                address_line_2 = safe_str(row.get(actual_columns.get('address_line_2', ''), ''))
                address_line_3 = safe_str(row.get(actual_columns.get('address_line_3', ''), ''))
                town = safe_str(row.get(actual_columns.get('town', ''), ''))
                county = safe_str(row.get(actual_columns.get('county', ''), ''))
                postcode = safe_str(row.get(actual_columns.get('postcode', ''), ''))

                address_parts = [p for p in [address_line_1, address_line_2, address_line_3, town, county] if p and p.lower() != 'nan']
                address = ', '.join(address_parts)
                site_address = site_name or address

                # MPAN fields
                mpan_top = safe_str(row.get(actual_columns.get('mpan_top', ''), ''))
                mpan_bottom = safe_str(row.get(actual_columns.get('mpan_bottom', ''), ''))

                # Contract fields
                supplier_name = safe_str(row.get(actual_columns.get('supplier', ''), ''))
                old_supplier_name = safe_str(row.get(actual_columns.get('old_supplier', ''), ''))
                payment_type = safe_str(row.get(actual_columns.get('payment_type', ''), ''))
                annual_usage = parse_number(row.get(actual_columns.get('annual_usage', '')))
                start_date = parse_date(row.get(actual_columns.get('start_date', '')))
                end_date = parse_date(row.get(actual_columns.get('contract_end', '')))
                stand_charge = parse_number(row.get(actual_columns.get('stand_charge', '')))
                rate_1 = parse_number(row.get(actual_columns.get('rate_1', '')))
                net_notch = parse_number(row.get(actual_columns.get('net_notch', '')))
                rate_2 = parse_number(row.get(actual_columns.get('rate_2', '')))
                rate_3 = parse_number(row.get(actual_columns.get('rate_3', '')))
                comms_paid = parse_number(row.get(actual_columns.get('comms_paid', '')))
                company_number = safe_str(row.get(actual_columns.get('company_number', ''), ''))
                date_of_birth = parse_date(row.get(actual_columns.get('date_of_birth', '')))
                charity_ltd_company_number = safe_str(row.get(actual_columns.get('charity_ltd_company_number', ''), ''))
                month_sold = safe_str(row.get(actual_columns.get('month_sold', ''), ''))
                house_name = safe_str(row.get(actual_columns.get('house_name', ''), ''))
                house_number = safe_str(row.get(actual_columns.get('house_number', ''), ''))
                door_number = safe_str(row.get(actual_columns.get('door_number', ''), ''))
                term_sold = parse_number(row.get(actual_columns.get('term_sold', '')))
                aggregator = safe_str(row.get(actual_columns.get('aggregator', ''), ''))
                partner_details = safe_str(row.get(actual_columns.get('partner_details', ''), ''))
                bank_name = safe_str(row.get(actual_columns.get('bank_name', ''), ''))
                account_number = safe_str(row.get(actual_columns.get('ac_number', ''), ''))
                sort_code = safe_str(row.get(actual_columns.get('sort_code', ''), ''))
                home_door_number = safe_str(row.get(actual_columns.get('home_door_number', ''), ''))
                home_street = safe_str(row.get(actual_columns.get('home_street', ''), ''))
                partner_dob = parse_date(row.get(actual_columns.get('partner_dob', '')))
                credit_score = parse_number(row.get(actual_columns.get('credit_score', '')))
                data_source = safe_str(row.get(actual_columns.get('data_source', ''), ''))
                agent_sold = safe_str(row.get(actual_columns.get('agent_sold', ''), ''))


                # Get or create supplier
                supplier_id = None
                if supplier_name:
                    supplier_key = supplier_name.lower().strip()
                    supplier_id = suppliers_dict.get(supplier_key)
                    
                    if not supplier_id:
                        try:
                            new_supplier = Supplier_Master(
                                supplier_company_name=supplier_name,
                                supplier_contact_name='Auto-imported',
                                supplier_provisions=3,
                                created_at=datetime.utcnow()
                            )
                            session.add(new_supplier)
                            session.flush()
                            
                            supplier_id = new_supplier.supplier_id
                            suppliers_dict[supplier_key] = supplier_id
                            print(f"✨ Row {index + 2}: Created new supplier '{supplier_name}' (ID: {supplier_id})")
                            
                        except Exception as e:
                            print(f"❌ Row {index + 2}: Failed to create supplier '{supplier_name}': {e}")
                            supplier_id = 1
                else:
                    supplier_id = 1

                old_supplier_id = None
                if old_supplier_name:
                    old_supplier_key = old_supplier_name.lower().strip()
                    old_supplier_id = suppliers_dict.get(old_supplier_key)
                    
                    if not old_supplier_id:
                        try:
                            new_old_supplier = Supplier_Master(
                                supplier_company_name=old_supplier_name,
                                supplier_contact_name='Auto-imported',
                                supplier_provisions=3,
                                created_at=datetime.utcnow()
                            )
                            session.add(new_old_supplier)
                            session.flush()
                            old_supplier_id = new_old_supplier.supplier_id
                            suppliers_dict[old_supplier_key] = old_supplier_id
                        except Exception as e:
                            print(f"Failed to create old supplier: {e}")

                business_name = trading_name or client_name
                contact_person = main_contact or client_name
                phone = tel_no or mobile_no

                # Skip empty rows
                if not business_name and not phone and not email and not mpan_top and not contact_person:
                    continue
                
                # ✅ CHECK DUPLICATES WITH ARCHIVING
                if mpan_top:
                    mpan_key = mpan_top.strip().lower()
                    existing_records = existing_mpans.get(mpan_key)  # ✅ Changed from existing_contract to existing_records
                    
                    if existing_records:  # ✅ This is now a LIST
                        duplicate_count += 1
                        
                        # ✅ Check if ANY of the existing records belong to a different tenant
                        cross_tenant_record = None
                        same_tenant_record = None
                        
                        for record in existing_records:  # ✅ NOW we iterate through the list
                            if record['tenant_id'] != tenant_id:
                                # Found a cross-tenant duplicate - PRIORITY 1
                                cross_tenant_record = record
                                break  # Stop at first cross-tenant match
                            else:
                                # Found a same-tenant duplicate
                                same_tenant_record = record
                        
                        # ✅ PRIORITY 1: Cross-tenant duplicates are ALWAYS skipped
                        if cross_tenant_record:
                            cross_tenant_duplicates.append({
                                'row': index + 2,
                                'mpan': mpan_top,
                                'new_company': business_name,
                                'existing_company': cross_tenant_record['company_name'],
                                'existing_tenant_id': cross_tenant_record['tenant_id'],
                                'assigned_to': cross_tenant_record['assigned_to_name'],
                                'is_archived': cross_tenant_record['is_archived']
                            })
                            
                            print(f"⚠️ Row {index + 2}: CROSS-TENANT DUPLICATE FOUND!")
                            print(f"   MPAN: {mpan_top}")
                            print(f"   Your tenant: {tenant_id} | Existing tenant: {cross_tenant_record['tenant_id']}")
                            print(f"   Assigned to: {cross_tenant_record['assigned_to_name']}")
                            print(f"   ⏭️ Skipping import for this record")
                            continue
                        
                        # ✅ PRIORITY 2: Same-tenant duplicate handling (if no cross-tenant found)
                        if same_tenant_record:
                            existing_contract = same_tenant_record['contract']
                            existing_assigned_to = same_tenant_record['assigned_to_name']
                            is_archived = same_tenant_record['is_archived']
                            
                            # ✅ Get existing end date early
                            existing_end_date = existing_contract.contract_end_date
                            new_end_date = end_date
                            
                            project = session.query(Project_Details).filter_by(
                                project_id=existing_contract.project_id
                            ).first()
                            
                            if not project:
                                session.rollback()
                                error_count += 1
                                errors.append(f"Row {index + 2}: Project not found for MPAN {mpan_top}")
                                continue
                            
                            client = session.query(Client_Master).filter_by(
                                client_id=project.client_id
                            ).first()
                            
                            if not client:
                                session.rollback()
                                error_count += 1
                                errors.append(f"Row {index + 2}: Client not found for MPAN {mpan_top}")
                                continue
                            
                            # ✅ Determine the action for tracking
                            if existing_end_date and new_end_date and existing_end_date == new_end_date:
                                action = 'Exact duplicate - skipped'
                            elif existing_end_date and new_end_date and new_end_date < existing_end_date:
                                action = 'Older record - created as archived'
                            elif existing_end_date and new_end_date and new_end_date > existing_end_date:
                                action = 'Newer record - archived existing'
                            else:
                                action = 'Updated existing record'
                            
                            # Track same-tenant duplicate
                            duplicate_details.append({
                                'row': index + 2,
                                'mpan': mpan_top,
                                'company': business_name,
                                'assigned_to': existing_assigned_to,
                                'action': action
                            })
                            
                            # ✅ Skip if already archived
                            if is_archived:
                                print(f"⏭️ Row {index + 2}: Skipping - existing record already archived for MPAN {mpan_top}")
                                continue
                            
                            if not new_end_date:
                                print(f"⏭️ Row {index + 2}: Skipping - no end date in new record for MPAN {mpan_top}")
                                continue
                            
                            # ✅ If new record is OLDER, create it as ARCHIVED
                            if existing_end_date and new_end_date < existing_end_date:
                                print(f"⏭️ Row {index + 2}: Older record detected - creating as archived for MPAN {mpan_top}")
                                
                                try:
                                    # Create the older record as ARCHIVED
                                    archived_client = Client_Master(
                                        tenant_id=tenant_id,
                                        assigned_employee_id=opportunity_owner_id,
                                        client_company_name=business_name or '',
                                        client_contact_name=contact_person or '',
                                        address=address or '',
                                        post_code=postcode or '',
                                        client_phone=tel_no or '',
                                        client_mobile=mobile_no or None,
                                        client_email=email or '',
                                        client_website='',
                                        default_currency_id=1,
                                        created_at=datetime.utcnow(),
                                        position=position or None,
                                        company_number=company_number or None,
                                        date_of_birth=date_of_birth,
                                        charity_ltd_company_number=charity_ltd_company_number or None,
                                        partner_details=partner_details or None,
                                        bank_name=bank_name or None,
                                        account_number=account_number or None,
                                        sort_code=sort_code or None,
                                        is_archived=True,
                                        home_door_number=home_door_number or None,
                                        home_street=home_street or None,
                                        archived_at=datetime.utcnow(),
                                        archived_reason=f"Historical record (ended {new_end_date}) - superseded by existing contract ending {existing_end_date}",
                                        partner_dob=partner_dob,
                                        credit_score=credit_score,
                                    )
                                    session.add(archived_client)
                                    session.flush()
                                    
                                    archived_client_id = archived_client.client_id
                                    
                                    # Create archived project
                                    archived_project = Project_Details(
                                        client_id=archived_client_id,
                                        opportunity_id=None,
                                        project_title=business_name or '',
                                        project_description='Imported site location',
                                        start_date=start_date if start_date else datetime.utcnow().date(),
                                        end_date=end_date,
                                        employee_id=employee_id,
                                        assigned_employee_id=opportunity_owner_id,    
                                        status=None,                                  
                                        created_at=datetime.utcnow(),
                                        updated_at=datetime.utcnow(),
                                        address=site_address or address or '',
                                        Misc_Col1=None,
                                        Misc_Col2=int(annual_usage) if annual_usage else None,
                                        site_name=site_name or None,
                                        month_sold=month_sold or None,
                                        house_name=house_name or None,
                                        house_number=house_number or None,
                                        door_number=door_number or None,
                                        town=town or None,
                                        county=county or None,
                                    )
                                    session.add(archived_project)
                                    session.flush()
                                    
                                    # Create archived contract
                                    archived_contract = Energy_Contract_Master(
                                        project_id=archived_project.project_id,
                                        employee_id=employee_id,
                                        supplier_id=supplier_id or 1,
                                        old_supplier_id=old_supplier_id,
                                        contract_start_date=start_date or datetime.utcnow().date(),
                                        contract_end_date=end_date,
                                        terms_of_sale='',
                                        service_id=import_service_id,
                                        unit_rate=rate_1 or 0.0,
                                        currency_id=1,
                                        document_details=None,
                                        created_at=datetime.utcnow(),
                                        updated_at=datetime.utcnow(),
                                        mpan_number=mpan_top or '',
                                        mpan_bottom=mpan_bottom or '',
                                        net_notch=net_notch,
                                        term_sold=term_sold,
                                        rate_2=rate_2,
                                        rate_3=rate_3,
                                        comms_paid=comms_paid,
                                        standing_charge=stand_charge,
                                        aggregator=aggregator or None,
                                        rate_1=rate_1,
                                    )
                                    session.add(archived_contract)
                                    session.flush()
                                    
                                    # ✅ ADD TO existing_mpans so subsequent imports can find this archived record
                                    if mpan_key not in existing_mpans:
                                        existing_mpans[mpan_key] = []
                                    
                                    existing_mpans[mpan_key].append({
                                        'contract': archived_contract,
                                        'tenant_id': tenant_id,
                                        'assigned_to_id': opportunity_owner_id,
                                        'assigned_to_name': assigned_employee_name or 'Unassigned',
                                        'company_name': business_name,
                                        'is_archived': True  # This is an archived record
                                    })
                                    
                                    session.commit()
                                    success_count += 1
                                    print(f"✅ OLDER RECORD CREATED AS ARCHIVED")
                                    continue
                                    
                                except Exception as archive_error:
                                    session.rollback()
                                    print(f"❌ ARCHIVE CREATION FAILED for row {index + 2}: {archive_error}")
                                    import traceback
                                    traceback.print_exc()
                                    error_count += 1
                                    errors.append(f"Row {index + 2}: Archive creation failed - {str(archive_error)}")
                                    continue
                            
                            # ✅ Skip exact duplicates (same MPAN + same end date)
                            if existing_end_date and new_end_date == existing_end_date:
                                print(f"⏭️ Row {index + 2}: Skipping - exact duplicate (same MPAN and end date)")
                                continue
                            
                            # ✅ If new record is newer, archive existing
                            if existing_end_date and new_end_date > existing_end_date:
                                print(f"🔄 Row {index + 2}: New record is newer - archiving existing record for MPAN {mpan_top}")
                                
                                try:
                                    client.is_archived = True
                                    client.archived_at = datetime.utcnow()
                                    client.archived_reason = f"Superseded by newer contract (ending {new_end_date})"
                                    session.flush()
                                    print(f"✅ Archived existing client ID {client.client_id}")
                                    
                                    # ✅ UPDATE the existing_mpans entry to reflect archived status
                                    for record in existing_mpans[mpan_key]:
                                        if record['contract'].energy_contract_master_id == existing_contract.energy_contract_master_id:
                                            record['is_archived'] = True
                                            break
                                    
                                except Exception as archive_error:
                                    session.rollback()
                                    print(f"❌ Failed to archive existing record: {archive_error}")
                                    error_count += 1
                                    errors.append(f"Row {index + 2}: Failed to archive existing - {str(archive_error)}")
                                    continue

                                print(f"✅ Proceeding to create newer replacement record for MPAN {mpan_top}")

                # CREATE NEW CLIENT
                try:
                    new_client = Client_Master(
                        tenant_id=tenant_id,
                        assigned_employee_id=opportunity_owner_id,
                        client_company_name=business_name or '',  
                        client_contact_name=contact_person or '',  
                        address=address or '',
                        post_code=postcode or '',
                        home_door_number=home_door_number or None,
                        home_street=home_street or None,
                        client_phone=tel_no or '',
                        client_mobile=mobile_no or None,
                        client_email=email or '',
                        client_website='',
                        default_currency_id=1,
                        created_at=datetime.utcnow(),
                        position=position or None,
                        company_number=company_number or None,
                        date_of_birth=date_of_birth,
                        charity_ltd_company_number=charity_ltd_company_number or None,
                        partner_details=partner_details or None,
                        bank_name=bank_name or None,
                        account_number=account_number or None,
                        sort_code=sort_code or None,
                        partner_dob=partner_dob,
                        credit_score=credit_score,
                        is_archived=False,
                    )
                    session.add(new_client)
                    session.flush()
                    
                    client_id = new_client.client_id
                                        
                    # Create Project
                    project = None
                    if site_address or annual_usage or mpan_top or start_date or end_date:
                        project_start_date = start_date if start_date else datetime.utcnow().date()
                        project_end_date = end_date if end_date else None
                        project = Project_Details(
                            client_id=client_id,
                            opportunity_id=None,  # Renewals have no opportunity_id
                            project_title=business_name or 'Renewal Contract',
                            project_description='Imported renewal contract',
                            start_date=start_date if start_date else datetime.utcnow().date(),
                            end_date=end_date,
                            employee_id=employee_id,
                            assigned_employee_id=opportunity_owner_id,
                            status=None,
                            created_at=datetime.utcnow(),
                            updated_at=datetime.utcnow(),
                            address=site_address or address or '',
                            Misc_Col1=None,
                            Misc_Col2=int(annual_usage) if annual_usage else None,
                            site_name=site_name or None,
                            month_sold=month_sold or None,
                            house_name=house_name or None,
                            house_number=house_number or None,
                            door_number=door_number or None,
                            town=town or None,
                            county=county or None,
                        )
                        session.add(project)
                        session.flush()
                    
                    # Create Contract
                    if project and mpan_top:
                        from datetime import timedelta
                        
                        contract_start_date = start_date if start_date else datetime.utcnow().date()
                        
                        if end_date:
                            contract_end_date = end_date
                        else:
                            contract_end_date = contract_start_date + timedelta(days=365)
                        
                        contract = Energy_Contract_Master(
                            project_id=project.project_id,
                            employee_id=employee_id,
                            supplier_id=supplier_id or 1,
                            old_supplier_id=old_supplier_id,
                            contract_start_date=contract_start_date,
                            contract_end_date=contract_end_date,
                            terms_of_sale='',
                            service_id=import_service_id,
                            unit_rate=rate_1 or 0.0,
                            currency_id=1,
                            document_details=None,
                            created_at=datetime.utcnow(),
                            updated_at=datetime.utcnow(),
                            mpan_number=mpan_top or '',
                            mpan_bottom=mpan_bottom or '',
                            net_notch=net_notch,
                            rate_2=rate_2,
                            rate_3=rate_3,
                            comms_paid=comms_paid,
                            standing_charge=stand_charge,
                            aggregator=aggregator or None,
                            rate_1=rate_1,
                            payment_type=payment_type or None,
                        )
                        session.add(contract)
                        session.flush()

                        # ✅ Add new contract to existing_mpans as a LIST item
                        if mpan_top:
                            mpan_key = mpan_top.strip().lower()
                            
                            if mpan_key not in existing_mpans:
                                existing_mpans[mpan_key] = []
                            
                            existing_mpans[mpan_key].append({
                                'contract': contract,
                                'tenant_id': tenant_id,
                                'assigned_to_id': opportunity_owner_id,
                                'assigned_to_name': assigned_employee_name or 'Unassigned',
                                'company_name': business_name,
                                'is_archived': False
                            })
                                            
                        success_count += 1
                    
                    if (success_count + duplicate_count) % BATCH_SIZE == 0:
                        session.commit()
                        print(f"📊 Batch committed: {success_count + duplicate_count}/{total_rows}")
                
                # ✅ IMPROVED ERROR HANDLING - No more raw SQL errors!
                except Exception as client_error:
                    session.rollback()
                    error_count += 1
                    
                    error_str = str(client_error)
                    
                    # Handle duplicate key violations
                    if "UniqueViolation" in error_str or "duplicate key" in error_str:
                        error_msg = f"Customer '{contact_person}'"
                        if business_name:
                            error_msg += f" from '{business_name}'"
                        error_msg += " already exists (duplicate record)"
                        
                        errors.append(error_msg)
                        print(f"⚠️ Row {index + 2}: {error_msg}")
                    
                    # Handle integrity errors (missing required fields)
                    elif "IntegrityError" in error_str or "violates not-null constraint" in error_str:
                        errors.append(f"Missing required field - check your data")
                        print(f"⚠️ Row {index + 2}: Data validation error")
                    
                    # Handle foreign key errors
                    elif "ForeignKeyViolation" in error_str or "foreign key constraint" in error_str:
                        errors.append(f"Invalid reference (supplier, employee, etc.)")
                        print(f"⚠️ Row {index + 2}: Foreign key error")
                    
                    # Generic error - show simplified message
                    else:
                        error_lines = error_str.split('\n')
                        error_msg = error_lines[0] if error_lines else error_str
                        if len(error_msg) > 150:
                            error_msg = error_msg[:150] + "..."
                        
                        errors.append(error_msg)
                        print(f"❌ Row {index + 2}: {error_msg}")
                    
                    continue
                
            except Exception as row_error:
                session.rollback()
                error_count += 1
                
                error_str = str(row_error)
                
                if "UniqueViolation" in error_str or "duplicate key" in error_str:
                    error_msg = f"Duplicate customer detected"
                elif "IntegrityError" in error_str:
                    error_msg = f"Data validation error"
                else:
                    error_lines = error_str.split('\n')
                    error_msg = error_lines[0][:150] if error_lines else error_str[:150]
                
                errors.append(error_msg)
                print(f"❌ Row {index + 2}: {error_msg}")
                continue
                        
        # FINAL COMMIT
        try:
            session.commit()
            print(f"📊 Final batch committed: {success_count + duplicate_count}/{total_rows}")
        except Exception as commit_error:
            print(f"❌ Final commit error: {commit_error}")
            session.rollback()
        
        print(f"✅ Import complete: {success_count} new, {duplicate_count} updated, {error_count} errors")
        
        print(f"✅ Import complete: {success_count} new, {duplicate_count} duplicates, {error_count} errors")

        # ✅ BUILD DETAILED DUPLICATE REPORT
        duplicate_report = []
        if duplicate_details:
            duplicate_report.append("\n📋 SAME-TENANT DUPLICATES:")
            for dup in duplicate_details:
                duplicate_report.append(
                    f"  Row {dup['row']}: {dup['company']} (MPAN: {dup['mpan']}) - "
                    f"Assigned to: {dup['assigned_to']} - {dup['action']}"
                )

        if cross_tenant_duplicates:
            duplicate_report.append("\n⚠️ CROSS-TENANT DUPLICATES (SKIPPED):")
            for dup in cross_tenant_duplicates:
                archived_status = " [ARCHIVED]" if dup['is_archived'] else ""
                duplicate_report.append(
                    f"  Row {dup['row']}: {dup['new_company']} (MPAN: {dup['mpan']}) - "
                    f"Already exists in another account{archived_status} - "
                    f"Assigned to: {dup['assigned_to']}"
                )

        # ✅ ADD DEBUG PRINT TO SEE WHAT'S IN THE LISTS
        print(f"\n🔍 DUPLICATE REPORT DEBUG:")
        print(f"   duplicate_details count: {len(duplicate_details)}")
        print(f"   cross_tenant_duplicates count: {len(cross_tenant_duplicates)}")
        print(f"   duplicate_report lines: {len(duplicate_report)}")
        print(f"   duplicate_report content: {duplicate_report}")

        return jsonify({
            'success': True,
            'message': f'Import completed',
            'total_rows': len(df),
            'successful': success_count,
            'duplicates': duplicate_count,
            'same_tenant_duplicates': len(duplicate_details),
            'cross_tenant_duplicates': len(cross_tenant_duplicates),
            'failed': error_count,
            'errors': errors[:50],
            'duplicate_report': duplicate_report,  # ✅ Make sure this line exists!
            'assigned_to': assigned_employee_name,
            'assigned_employee_id': assigned_employee_id
        }), 200
        
    except Exception as e:
        session.rollback()
        print(f"\n\n❌❌❌ EXCEPTION CAUGHT ❌❌❌")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"❌❌❌ END EXCEPTION ❌❌❌\n\n")
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
    finally:
        sql_logger.setLevel(original_level)
        session.close()


@import_bp.route('/template', methods=['GET'])
@token_required
def download_template():
    """Download Excel template matching Cash2Switch format"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import send_file
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Renewals Import Template"
        
        # Headers (matching exact Excel structure)
        headers = [
            "Client Name", "Trading Name", "Main Contact", "Position", "Tel No", "Mobile No",
            "Email", "Site Name", "Month Sold", "", "Address Line 1", "Address Line 2",
            "Address Line 3", "Town", "County", "Postcode", "Mpan Top", "Mpan Bottom",
            "", "", "Data Source", "Welcome Call", "Payment Type", "Supplier", "Net Notch",
            "In Contract", "Agent Sold", "Start Date", "Contract End", "Stand Charge",
            "Rate 1", "Rate 2", "Rate 3", "", "", "Aggregator", "Annual Usage",
            "Comms Paid", "Company Number", "Date of Birth", "Bank Name", "Ac Number",
            "Sort Code", "Charity/Ltd Company Number", "Partner Details"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if header:  # Only style non-empty headers
                cell.fill = header_fill
                cell.font = header_font
        
        # Add example row
        example = [
            "ABC Limited",  # Client Name
            "ABC Trading",  # Trading Name
            "John Smith",   # Main Contact
            "Director",     # Position
            "07700900000",  # Tel No
            "07700900001",  # Mobile No
            "john@abc.com", # Email
            "Main Site",    # Site Name
            "Jan-24",       # Month Sold
            "",             # Empty
            "123 Main St",  # Address Line 1
            "Unit 5",       # Address Line 2
            "Industrial Estate",  # Address Line 3
            "London",       # Town
            "Greater London",  # County
            "SW1A 1AA",     # Postcode
            "1100012314490",  # Mpan Top
            "04031N12",     # Mpan Bottom
            "", "",         # Empty
            "Renewals",     # Data Source
            "Yes",          # Welcome Call
            "DD",           # Payment Type
            "British Gas",  # Supplier
            "0.1",          # Net Notch
            "1 Year",       # In Contract
            "Sales Team",   # Agent Sold
            "01/01/2024",   # Start Date
            "31/12/2024",   # Contract End
            "45.13",        # Stand Charge
            "35.00",        # Rate 1
            "26.46",        # Rate 2
            "",             # Rate 3
            "", "",         # Empty
            "Online",       # Aggregator
            "25000",        # Annual Usage
            "£7.92",        # Comms Paid
            "12345678",     # Company Number
            "",             # Date of Birth
            "Barclays",     # Bank Name
            "12345678",     # Ac Number
            "20-00-00",     # Sort Code
            "",             # Charity/Ltd Company Number
            ""              # Partner Details
        ]
        
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 30)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='cash2switch_renewals_template.xlsx'
        )
        
    except Exception as e:
        current_app.logger.exception(f"❌ Template download failed: {e}")
        return jsonify({'error': 'Failed to generate template'}), 500

@import_bp.route('/energy-clients/reset-sequence', methods=['POST'])
@jwt_required()
def reset_energy_client_sequence():
    """Reset the client_id sequence after deleting all customers"""
    current_user = get_jwt_identity()
    tenant_id = current_user.get('tenant_id')
    
    session = SessionLocal()  # ✅ Use SessionLocal instead of db.session
    
    try:
        # Reset Client_Master sequence
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence('"StreemLyne_MT"."Client_Master"', 'client_id'),
                COALESCE((SELECT MAX(client_id) FROM "StreemLyne_MT"."Client_Master" WHERE tenant_id = :tenant_id), 0),
                true
            )
        """), {'tenant_id': tenant_id})
        
        # Reset Project_Details sequence
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence('"StreemLyne_MT"."Project_Details"', 'project_id'),
                COALESCE((SELECT MAX(project_id) FROM "StreemLyne_MT"."Project_Details"), 0),
                true
            )
        """))
        
        # Reset Energy_Contract_Master sequence  
        session.execute(text("""
            SELECT setval(
                pg_get_serial_sequence('"StreemLyne_MT"."Energy_Contract_Master"', 'energy_contract_master_id'),
                COALESCE((SELECT MAX(energy_contract_master_id) FROM "StreemLyne_MT"."Energy_Contract_Master"), 0),
                true
            )
        """))
        
        session.commit()  # ✅ Changed from db.session.commit()
        
        return jsonify({
            'message': 'All sequences reset successfully',
            'success': True
        }), 200
        
    except Exception as e:
        session.rollback()  # ✅ Changed from db.session.rollback()
        logger.error(f"Error resetting sequences: {str(e)}")
        return jsonify({'error': 'Failed to reset sequences'}), 500
    finally:
        session.close()  # ✅ Always close the session

def handle_duplicate_customer(session, tenant_id, mpan_top, phone, new_end_date, new_client_id=None):
    """
    Check for existing customer by MPAN Top or Phone.
    If found, compare end dates and archive the older one.
    Returns: (should_archive_new, existing_client_id)
    """
    from datetime import datetime
    
    # Find existing customer by MPAN Top or Phone
    existing_query = session.query(
        Client_Master.client_id,
        Client_Master.client_company_name,
        Energy_Contract_Master.contract_end_date,
        Energy_Contract_Master.mpan_number
    ).join(
        Project_Details, Client_Master.client_id == Project_Details.client_id
    ).join(
        Energy_Contract_Master, Project_Details.project_id == Energy_Contract_Master.project_id
    ).filter(
        Client_Master.tenant_id == tenant_id,
        Client_Master.is_deleted == False,
        Client_Master.is_archived == False,  # Only check non-archived records
        or_(
            Energy_Contract_Master.mpan_number == mpan_top,
            Client_Master.client_phone == phone
        )
    )
    
    # Exclude the current client if this is an update
    if new_client_id:
        existing_query = existing_query.filter(Client_Master.client_id != new_client_id)
    
    existing = existing_query.first()
    
    if not existing:
        return False, None  # No duplicate found
    
    existing_client_id, existing_name, existing_end_date, existing_mpan = existing
    
    # Compare end dates
    if not new_end_date and not existing_end_date:
        # Both have no end date - keep existing
        return True, existing_client_id
    
    if not new_end_date:
        # New has no end date - archive new, keep existing
        return True, existing_client_id
    
    if not existing_end_date:
        # Existing has no end date - archive existing, keep new
        return False, existing_client_id
    
    # Both have end dates - compare
    new_date = new_end_date if isinstance(new_end_date, datetime) else datetime.strptime(str(new_end_date), '%Y-%m-%d')
    existing_date = existing_end_date if isinstance(existing_end_date, datetime) else existing_end_date
    
    if new_date > existing_date:
        # New is more recent - archive existing
        current_app.logger.info(f"📦 Archiving older record: {existing_name} (End: {existing_end_date}) - New end date: {new_end_date}")
        return False, existing_client_id
    else:
        # Existing is more recent - archive new
        current_app.logger.info(f"📦 New record is older - will archive after creation (End: {new_end_date} vs {existing_end_date})")
        return True, existing_client_id


def archive_customer(session, client_id, reason="Superseded by newer contract"):
    """
    Archive a customer record
    """
    from datetime import datetime
    
    client = session.query(Client_Master).filter_by(client_id=client_id).first()
    if client:
        client.is_archived = True
        client.archived_at = datetime.utcnow()
        client.archived_reason = reason
        
        current_app.logger.info(f"📦 Archived client {client_id}: {client.client_company_name}")

@import_bp.route('/leads', methods=['POST', 'OPTIONS'])
@token_required
def import_leads():
    """
    ✅ OPTIMIZED: Handles 5,000-10,000 leads via bulk inserts
    """
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    filename = secure_filename(file.filename)
    if not ('.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls', 'csv'}):
        return jsonify({'error': 'Invalid file type'}), 400

    tenant_id = get_tenant_id_from_user(request.current_user)
    if not tenant_id:
        return jsonify({'error': 'Tenant not found'}), 400

    importing_employee_id = request.current_user.employee_id
    assigned_employee_id = request.form.get('assigned_employee_id', type=int)
    opportunity_owner_id = assigned_employee_id if assigned_employee_id else importing_employee_id

    session = SessionLocal()
    try:
        # Validate assigned employee
        assigned_employee_name = None
        if assigned_employee_id:
            ae = session.query(Employee_Master).filter_by(
                employee_id=assigned_employee_id,
                tenant_id=tenant_id
            ).first()
            if ae:
                assigned_employee_name = ae.employee_name
            else:
                return jsonify({'error': f'Invalid employee ID: {assigned_employee_id}'}), 400
        else:
            own_emp = session.query(Employee_Master).filter_by(
                employee_id=importing_employee_id,
                tenant_id=tenant_id
            ).first()
            assigned_employee_name = own_emp.employee_name if own_emp else None

        service_param = request.args.get('service', 'electricity')
        service_id_map = {'electricity': 1, 'utilities': 1, 'water': 2, 'gas': 3}
        import_service_id = service_id_map.get(service_param.strip().lower(), 1)

        # ── Read file ──────────────────────────────────────────────────
        file_ext = filename.rsplit('.', 1)[1].lower()
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name
            
            if file_ext == 'csv':
                df = pd.read_csv(tmp_path, encoding='utf-8-sig', dtype=str)
            else:
                try:
                    df = pd.read_excel(tmp_path, engine='openpyxl', dtype=str)
                except Exception:
                    df = pd.read_excel(tmp_path, engine='xlrd', dtype=str)
            
            os.unlink(tmp_path)
        except Exception as e:
            return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

        # ── Column mapping (same as before) ────────────────────────────
        original_cols = list(df.columns)
        df.columns = df.columns.str.strip().str.lower().str.replace('_', ' ').str.replace(r'\s+', ' ', regex=True)

        column_map = {
            'client_name':    ['client name', 'business name', 'company name'],
            'trading_name':   ['trading name', 'business', 'company'],
            'main_contact':   ['main contact', 'contact person', 'contact'],
            'tel_no':         ['tel no', 'phone', 'telephone', 'tel'],
            'mobile_no':      ['mobile no', 'mobile', 'cell'],
            'email':          ['email', 'e-mail'],
            'mpan_top':       ['mpan top', 'mpan core', 'mpan mpr', 'mpan', 'mpr'],
            'mpan_bottom':    ['mpan bottom', 'mpan llf'],
            'supplier':       ['supplier', 'supplier name'],
            'start_date':     ['start date', 'contract start'],
            'contract_end':   ['contract end', 'end date', 'expiry'],
            'annual_usage':   ['annual usage', 'usage', 'kwh'],
            'payment_type':   ['payment type'],
            'postcode':       ['postcode', 'post code', 'zip'],
        }

        actual_columns = {}
        for field, aliases in column_map.items():
            for col in df.columns:
                if col in aliases:
                    actual_columns[field] = col
                    break

        # Restore original columns
        df.columns = original_cols
        norm_to_orig = {}
        for orig in original_cols:
            import re
            normed = re.sub(r'\s+', ' ', orig.strip().lower().replace('_', ' '))
            norm_to_orig[normed] = orig

        actual_columns_orig = {}
        for field, normed_col in actual_columns.items():
            actual_columns_orig[field] = norm_to_orig.get(normed_col, normed_col)

        def gcol(field):
            return actual_columns_orig.get(field, '')

        # ── Get default stage ───────────────────────────────────────────
        from sqlalchemy import text as sa_text
        stage_row = session.execute(sa_text("""
            SELECT stage_id FROM "StreemLyne_MT"."Stage_Master"
            WHERE LOWER(stage_name) = 'not called' LIMIT 1
        """)).fetchone()
        
        if stage_row:
            default_stage_id = stage_row[0]
        else:
            from backend.models import Stage_Master
            first_stage = session.query(Stage_Master).order_by(Stage_Master.stage_id).first()
            default_stage_id = first_stage.stage_id if first_stage else 1

        # ✅ CRITICAL OPTIMIZATION: Load existing MPANs into a SET (not dict with lists)
        logger.info('🔍 Loading existing MPANs for duplicate checking...')
        existing_mpans_set = set()
        
        existing_leads_rows = session.execute(sa_text("""
            SELECT LOWER(TRIM(od."mpan_mpr")) as mpan_key
            FROM "StreemLyne_MT"."Opportunity_Details" od
            WHERE od."tenant_id" = :tenant_id
              AND od."mpan_mpr" IS NOT NULL 
              AND od."mpan_mpr" != ''
        """), {'tenant_id': tenant_id}).fetchall()
        
        for row in existing_leads_rows:
            existing_mpans_set.add(row[0])
        
        logger.info(f'✅ Loaded {len(existing_mpans_set)} existing MPANs')

        # ✅ BULK PROCESSING: Prepare all rows first, then insert in batches
        total_rows = len(df)
        success_count = 0
        duplicate_count = 0
        error_count = 0
        errors = []
        
        # Store validated rows for bulk insert
        bulk_insert_data = []
        duplicate_report = []

        logger.info(f'📊 Processing {total_rows} rows...')

        for index, row in df.iterrows():
            try:
                # Extract data (same as before)
                client_name = safe_str(row.get(gcol('client_name'), ''))
                trading_name = safe_str(row.get(gcol('trading_name'), ''))
                main_contact = safe_str(row.get(gcol('main_contact'), ''))
                tel_no = safe_str(row.get(gcol('tel_no'), ''))
                mobile_no = safe_str(row.get(gcol('mobile_no'), ''))
                email = safe_str(row.get(gcol('email'), ''))
                mpan_top = safe_str(row.get(gcol('mpan_top'), ''))
                mpan_bottom = safe_str(row.get(gcol('mpan_bottom'), ''))
                supplier_name = safe_str(row.get(gcol('supplier'), ''))
                start_date = parse_date(row.get(gcol('start_date'), ''))
                end_date = parse_date(row.get(gcol('contract_end'), ''))
                annual_usage = parse_number(row.get(gcol('annual_usage'), ''))
                payment_type = safe_str(row.get(gcol('payment_type'), ''))
                postcode = safe_str(row.get(gcol('postcode'), ''))

                business_name = trading_name or client_name
                contact_person = main_contact or client_name
                phone = tel_no or mobile_no

                # Skip empty rows
                if not business_name and not phone and not email and not mpan_top:
                    continue

                # ✅ FAST DUPLICATE CHECK (set lookup is O(1) vs O(n) list iteration)
                if mpan_top:
                    mpan_key = mpan_top.strip().lower()
                    
                    if mpan_key in existing_mpans_set:
                        duplicate_count += 1
                        duplicate_report.append(f"Row {index + 2}: {business_name} (MPAN: {mpan_top}) - Duplicate")
                        continue

                # Find supplier_id
                supplier_id = None
                if supplier_name:
                    sup = session.query(Supplier_Master).filter(
                        Supplier_Master.supplier_company_name.ilike(f'%{supplier_name}%')
                    ).first()
                    if sup:
                        supplier_id = sup.supplier_id

                # ✅ ADD TO BULK INSERT LIST (no database writes yet)
                bulk_insert_data.append({
                    'tenant_id': tenant_id,
                    'title': business_name or contact_person or f'Lead {index + 2}',
                    'opp_date': datetime.utcnow().date(),
                    'owner_id': opportunity_owner_id,
                    'stage_id': default_stage_id,
                    'created_at': datetime.utcnow(),
                    'business_name': business_name or None,
                    'contact_person': contact_person or None,
                    'tel_number': phone or None,
                    'mobile_no': mobile_no or None,
                    'email': email or None,
                    'mpan_mpr': mpan_top or None,
                    'mpan_bottom': mpan_bottom or None,
                    'start_date': start_date,
                    'end_date': end_date,
                    'service_id': import_service_id,
                    'supplier_id': supplier_id,
                    'annual_usage': int(annual_usage) if annual_usage else None,
                    'payment_type': payment_type or None,
                    'postcode': postcode or None,
                })

                # Track this MPAN to prevent within-file duplicates
                if mpan_top:
                    existing_mpans_set.add(mpan_top.strip().lower())

            except Exception as row_err:
                error_count += 1
                err_str = str(row_err).split('\n')[0][:150]
                errors.append(f"Row {index + 2}: {err_str}")
                continue

        # ✅ BULK INSERT: Insert all rows in batches of 1000
        BATCH_SIZE = 1000
        total_to_insert = len(bulk_insert_data)
        
        logger.info(f'✅ Validation complete: {total_to_insert} rows ready for insert')
        logger.info(f'📥 Starting bulk insert in batches of {BATCH_SIZE}...')

        for batch_start in range(0, total_to_insert, BATCH_SIZE):
            batch = bulk_insert_data[batch_start:batch_start + BATCH_SIZE]
            
            try:
                # Use SQLAlchemy bulk_insert_mappings for maximum performance
                from backend.models import Opportunity_Details
                
                session.bulk_insert_mappings(Opportunity_Details, [
                    {
                        'tenant_id': row['tenant_id'],
                        'client_id': None,
                        'opportunity_title': row['title'],
                        'opportunity_description': 'Imported lead',
                        'opportunity_date': row['opp_date'],
                        'opportunity_owner_employee_id': row['owner_id'],
                        'stage_id': row['stage_id'],
                        'opportunity_value': 0,
                        'currency_id': 1,
                        'created_at': row['created_at'],
                        'business_name': row['business_name'],
                        'contact_person': row['contact_person'],
                        'tel_number': row['tel_number'],
                        'mobile_no': row['mobile_no'],
                        'email': row['email'],
                        'mpan_mpr': row['mpan_mpr'],
                        'mpan_bottom': row['mpan_bottom'],
                        'start_date': row['start_date'],
                        'end_date': row['end_date'],
                        'service_id': row['service_id'],
                        'supplier_id': row['supplier_id'],
                        'annual_usage': row['annual_usage'],
                        'payment_type': row['payment_type'],
                        'postcode': row['postcode'],
                        'is_allocated': False,
                    }
                    for row in batch
                ])
                
                session.flush()
                success_count += len(batch)
                
                logger.info(f'✅ Inserted batch {batch_start}-{batch_start + len(batch)}/{total_to_insert}')
                
            except Exception as batch_err:
                session.rollback()
                logger.error(f'❌ Batch insert failed: {batch_err}')
                error_count += len(batch)
                errors.append(f'Batch {batch_start}-{batch_start + len(batch)}: {str(batch_err)[:150]}')
                continue

        # ✅ FINAL COMMIT
        try:
            session.commit()
            logger.info(f'✅ Final commit successful: {success_count} leads inserted')
        except Exception as commit_err:
            session.rollback()
            logger.error(f'❌ Final commit failed: {commit_err}')
            return jsonify({'error': f'Commit failed: {str(commit_err)}'}), 500

        # ✅ RECALCULATE DISPLAY_ORDER (only once at the end)
        from backend.routes.crm_routes import recalculate_lead_display_order
        try:
            recalculate_lead_display_order(session, tenant_id, opportunity_owner_id)
            session.commit()
            logger.info(f'✅ Recalculated display_order for employee {opportunity_owner_id}')
        except Exception as recalc_err:
            logger.error(f'❌ Display order recalculation failed: {recalc_err}')

        return jsonify({
            'success': True,
            'message': f'Imported {success_count} leads',
            'total_rows': total_rows,
            'successful': success_count,
            'duplicates': duplicate_count,
            'failed': error_count,
            'errors': errors[:50],
            'duplicate_report': duplicate_report[:50],
            'assigned_to': assigned_employee_name,
            'assigned_employee_id': opportunity_owner_id,
        }), 200

    except Exception as e:
        session.rollback()
        import traceback
        traceback.print_exc()
        logger.error(f'❌ Import failed: {str(e)}')
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
    finally:
        session.close()

@import_bp.route('/leads/template', methods=['GET'])
@token_required
def download_leads_template():
    return download_leads_template_handler()