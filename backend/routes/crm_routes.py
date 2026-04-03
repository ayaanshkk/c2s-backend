# -*- coding: utf-8 -*-
"""
CRM Routes Blueprint — uses SessionLocal (same connection as renewals).
All raw SQL queries use SQLAlchemy text() via SessionLocal so leads and
renewals always hit the same database connection.
"""
from flask import Blueprint, request, g, jsonify
from functools import wraps
from sqlalchemy import text
from backend.db import SessionLocal
from backend.crm.controllers.crm_controller import CRMController
from backend.crm.middleware.tenant_middleware import require_tenant
from .auth_helpers import token_required
import logging

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_admin_from_db(user) -> bool:
    """Look up role from DB — same pattern as energy_customer_routes.py."""
    session = SessionLocal()
    try:
        result = session.execute(text("""
            SELECT rm.role_name
            FROM "StreemLyne_MT"."User_Role_Mapping" urm
            JOIN "StreemLyne_MT"."Role_Master" rm ON urm.role_id = rm.role_id
            WHERE urm.user_id = :user_id
            LIMIT 1
        """), {'user_id': user.user_id}).fetchone()
        role = result[0] if result else None
        return role in ['Platform Admin', 'Tenant Super Admin']
    except Exception:
        return False
    finally:
        session.close()


def _serial(v):
    """Serialise a DB value to a JSON-safe Python type."""
    if v is None:
        return None
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return float(v)
    except ImportError:
        pass
    return v


def _rows_to_list(rows):
    """Convert SQLAlchemy MappingResult rows to serialised list of dicts."""
    if not rows:
        return []
    return [{k: _serial(v) for k, v in dict(row).items()} for row in rows]


def tenant_from_jwt(f):
    """Set g.tenant_id (as int) from request.current_user.tenant_id."""
    @wraps(f)
    def _wrap(*args, **kwargs):
        current_user = getattr(request, 'current_user', None)
        if not current_user or getattr(current_user, 'tenant_id', None) is None:
            return jsonify({'error': 'Missing tenant in token'}), 401
        try:
            g.tenant_id = int(getattr(current_user, 'tenant_id'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid tenant_id in token'}), 401
        return f(*args, **kwargs)
    return _wrap


# ── Blueprint ─────────────────────────────────────────────────────────────────

crm_bp = Blueprint('crm', __name__, url_prefix='/api/crm')
crm_controller = CRMController()


# ========================================
# LEAD ROUTES
# ========================================

# backend/routes/crm_routes.py

@crm_bp.route('/leads', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads():
    """
    Get all leads with team overview stats and per-employee display_order
    Admin sees all tenant leads, non-admin sees only their own non-allocated leads
    ✅ FIXED: Team stats now show correctly, is_allocated scoping fixed
    """
    import logging
    logger = logging.getLogger(__name__)
    
    session = SessionLocal()
    try:
        tenant_id    = g.tenant_id
        current_user = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1
        exclude_stage = request.args.get('exclude_stage', '')
        employee_id   = getattr(current_user, 'employee_id', None)

        logger.warning(
            '🔍 get_leads: employee_id=%s tenant=%s service=%s',
            employee_id, tenant_id, service_param
        )

        if not employee_id:
            logger.warning('⚠️ No employee_id - returning empty')
            return jsonify({'data': [], 'team_stats': []}), 200

        # ✅ Determine if admin
        is_admin = _is_admin_from_db(current_user)
        
        # ================================================================
        # 1. RECALCULATE DISPLAY_ORDER (per employee, starting from 1)
        # ================================================================
        try:
            session.execute(text("""
                UPDATE "StreemLyne_MT"."Opportunity_Details" od
                SET display_order = sub.rn
                FROM (
                    SELECT opportunity_id,
                           ROW_NUMBER() OVER (
                               PARTITION BY opportunity_owner_employee_id
                               ORDER BY created_at ASC
                           ) AS rn
                    FROM "StreemLyne_MT"."Opportunity_Details"
                    WHERE tenant_id = :tenant_id
                      AND (is_allocated = FALSE OR is_allocated IS NULL)
                      AND opportunity_owner_employee_id IS NOT NULL
                ) sub
                WHERE od.opportunity_id = sub.opportunity_id
                  AND od.tenant_id = :tenant_id
            """), {'tenant_id': tenant_id})
            session.commit()
            logger.warning('✅ Recalculated display_order for all employees')
        except Exception as e:
            session.rollback()
            logger.error(f'❌ Error recalculating display_order: {e}')

        # ================================================================
        # 2. TEAM STATS QUERY - ✅ FIXED: Removed HAVING clause
        # ================================================================
        if is_admin:
            # Admin: Show ALL employees in tenant with their lead counts (including 0)
            team_stats_rows = session.execute(text("""
                SELECT 
                    em."employee_id",
                    em."employee_name",
                    COUNT(od."opportunity_id") as lead_count
                FROM "StreemLyne_MT"."Employee_Master" em
                LEFT JOIN "StreemLyne_MT"."Opportunity_Details" od 
                    ON em."employee_id" = od."opportunity_owner_employee_id"
                    AND od."tenant_id" = :tenant_id
                    AND od."service_id" = :service_id
                    AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
                WHERE em."tenant_id" = :tenant_id
                GROUP BY em."employee_id", em."employee_name"
                ORDER BY em."employee_name"
            """), {'tenant_id': tenant_id, 'service_id': service_id}).mappings().all()
        else:
            # Non-admin: Show only own stats
            team_stats_rows = session.execute(text("""
                SELECT 
                    em."employee_id",
                    em."employee_name",
                    COUNT(od."opportunity_id") as lead_count
                FROM "StreemLyne_MT"."Employee_Master" em
                LEFT JOIN "StreemLyne_MT"."Opportunity_Details" od 
                    ON em."employee_id" = od."opportunity_owner_employee_id"
                    AND od."tenant_id" = :tenant_id
                    AND od."service_id" = :service_id
                    AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
                WHERE em."employee_id" = :employee_id
                GROUP BY em."employee_id", em."employee_name"
            """), {'tenant_id': tenant_id, 'service_id': service_id, 'employee_id': employee_id}).mappings().all()

        team_stats = [
            {
                'employee_id': row['employee_id'],
                'employee_name': row['employee_name'],
                'lead_count': int(row['lead_count'] or 0)
            }
            for row in team_stats_rows
        ]

        logger.warning('📊 Team stats: %s', team_stats)

        # ================================================================
        # 3. MAIN LEADS QUERY - ✅ Properly scoped by employee and is_allocated
        # ================================================================
        query = """
            SELECT
                od.*,
                od.display_order,
                sm."stage_name",
                em."employee_name" AS assigned_to_name,
                COALESCE(od."business_name", od."opportunity_title") AS business_name,
                sup."supplier_company_name" AS supplier_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od."stage_id"    = sm."stage_id"
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od."opportunity_owner_employee_id" = em."employee_id"
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od."supplier_id"  = sup."supplier_id"
            WHERE od."tenant_id" = :tenant_id
            AND od."service_id" = :service_id
            AND od."opportunity_owner_employee_id" = :employee_id
            AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
        """
        params = {'tenant_id': tenant_id, 'service_id': service_id, 'employee_id': employee_id}
        # else:
        #     # Non-admin sees only their own non-allocated leads
        #     query = """
        #         SELECT
        #             od.*,
        #             od.display_order,
        #             sm."stage_name",
        #             em."employee_name" AS assigned_to_name,
        #             COALESCE(od."business_name", od."opportunity_title") AS business_name,
        #             sup."supplier_company_name" AS supplier_name
        #         FROM "StreemLyne_MT"."Opportunity_Details" od
        #         LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od."stage_id"    = sm."stage_id"
        #         LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od."opportunity_owner_employee_id" = em."employee_id"
        #         LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od."supplier_id"  = sup."supplier_id"
        #         WHERE od."tenant_id" = :tenant_id
        #           AND od."service_id" = :service_id
        #           AND od."opportunity_owner_employee_id" = :employee_id
        #           AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
        #     """
        #     params = {'tenant_id': tenant_id, 'service_id': service_id, 'employee_id': employee_id}

        if exclude_stage:
            query += ' AND (sm."stage_name" IS NULL OR LOWER(sm."stage_name") != LOWER(:exclude_stage))'
            params['exclude_stage'] = exclude_stage

        query += ' ORDER BY od."display_order" ASC'

        rows = session.execute(text(query), params).mappings().all()
        results = _rows_to_list(rows)

        logger.warning(
            '✅ get_leads returning %d leads + %d team stats for employee_id=%s (is_admin=%s)',
            len(results), len(team_stats), employee_id, is_admin
        )

        return jsonify({
            'data': results,
            'team_stats': team_stats,
            'user_context': {
                'is_admin': is_admin,
                'employee_id': employee_id
            }
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error('❌ get_leads error: %s', str(e))
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@crm_bp.route('/leads/<int:opportunity_id>', methods=['GET'])
@token_required
@tenant_from_jwt
def get_lead_detail(opportunity_id):
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id

        sql = """
            SELECT
                od.*,
                sm.stage_name,
                em.employee_name          AS assigned_to_name,
                COALESCE(od.business_name, od.opportunity_title) AS business_name,
                sup.supplier_company_name AS supplier_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id   = sm.stage_id
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
            WHERE od.tenant_id = :tenant_id
        """

        row = session.execute(
            text(sql + ' AND od.tenant_lead_id = :id LIMIT 1'),
            {'tenant_id': tenant_id, 'id': opportunity_id}
        ).mappings().first()

        if not row:
            row = session.execute(
                text(sql + ' AND od.opportunity_id = :id LIMIT 1'),
                {'tenant_id': tenant_id, 'id': opportunity_id}
            ).mappings().first()

        if not row:
            return jsonify({'error': 'Lead not found'}), 404

        return jsonify({k: _serial(v) for k, v in dict(row).items()}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


@crm_bp.route('/leads', methods=['POST'])
@token_required
@tenant_from_jwt
def create_lead():
    return crm_controller.create_lead()


@crm_bp.route('/leads/<int:opportunity_id>', methods=['PUT', 'PATCH'])
@token_required
@tenant_from_jwt
def update_lead(opportunity_id):
    if request.method == 'PATCH':
        ALLOWED = {
            'stage_id','status','business_name','contact_person','tel_number',
            'mobile_no','email','position','company_number','date_of_birth',
            'opportunity_owner_employee_id','mpan_mpr','mpan_bottom','supplier_id',
            'annual_usage','start_date','end_date','payment_type','term_sold',
            'net_notch','comms_paid','aggregator','site_name','month_sold',
            'house_name','house_number','door_number','address','town','county',
            'postcode','stand_charge','rate_1','rate_2','rate_3','night_charge',
            'eve_weekend_charge','other_charges_1','other_charges_2','other_charges_3',
            'bank_name','bank_account_number','bank_sort_code',
            'charity_ltd_company_number','partner_details',
            'meter_ref','uplift','comments','document_details',
        }
        session = SessionLocal()
        try:
            tenant_id = g.tenant_id
            data = request.get_json() or {}
            fields = {k: v for k, v in data.items() if k in ALLOWED}
            if not fields:
                return jsonify({'error': 'No valid fields provided'}), 400

            id_row = session.execute(text("""
                SELECT opportunity_id FROM "StreemLyne_MT"."Opportunity_Details"
                WHERE tenant_id = :t
                AND (tenant_lead_id = :id OR opportunity_id = :id)
                LIMIT 1
            """), {'t': tenant_id, 'id': opportunity_id}).mappings().first()

            if not id_row:
                return jsonify({'error': 'Lead not found'}), 404

            real_id = id_row['opportunity_id']
            set_clause = ', '.join(f'"{k}" = :{k}' for k in fields)
            params = {**fields, 'real_id': real_id, 'tenant_id': tenant_id}

            session.execute(text(
                f'UPDATE "StreemLyne_MT"."Opportunity_Details" '
                f'SET {set_clause} '
                f'WHERE opportunity_id = :real_id AND tenant_id = :tenant_id'
            ), params)
            session.commit()

            updated = session.execute(text("""
                SELECT od.*, sm.stage_name,
                       em.employee_name AS assigned_to_name,
                       COALESCE(od.business_name, od.opportunity_title) AS business_name,
                       sup.supplier_company_name AS supplier_name
                FROM "StreemLyne_MT"."Opportunity_Details" od
                LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id   = sm.stage_id
                LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
                LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
                WHERE od.opportunity_id = :id AND od.tenant_id = :t LIMIT 1
            """), {'id': real_id, 't': tenant_id}).mappings().first()

            return jsonify({k: _serial(v) for k, v in dict(updated).items()} if updated else {'success': True}), 200

        except Exception as e:
            session.rollback()
            import traceback; traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            try: session.close()
            except Exception: pass

    return crm_controller.update_lead(opportunity_id)


@crm_bp.route('/leads/<int:opportunity_id>/status', methods=['PATCH'])
@token_required
@tenant_from_jwt
def update_lead_status(opportunity_id):
    return crm_controller.update_lead_status(opportunity_id)


@crm_bp.route('/leads/assign', methods=['PATCH'])
@token_required
@tenant_from_jwt
def assign_leads():
    return crm_controller.assign_leads()


@crm_bp.route('/leads/<int:opportunity_id>', methods=['DELETE'])
@token_required
@tenant_from_jwt
def delete_lead(opportunity_id):
    return crm_controller.delete_lead(opportunity_id)


@crm_bp.route('/leads/search-all', methods=['GET'])
@token_required
@tenant_from_jwt
def search_all_leads():
    session = SessionLocal()
    try:
        tenant_id     = g.tenant_id
        q             = request.args.get('q', '').strip()
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1

        if not q or len(q) < 2:
            return jsonify([]), 200

        like = f'%{q}%'
        rows = session.execute(text("""
            SELECT od.*, sm.stage_name,
                   em.employee_name AS assigned_to_name,
                   COALESCE(od.business_name, od.opportunity_title) AS business_name,
                   sup.supplier_company_name AS supplier_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id   = sm.stage_id
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
            WHERE od.tenant_id  = :tenant_id
            AND   od.service_id = :service_id
            AND (
                COALESCE(od.business_name, od.opportunity_title) ILIKE :q
                OR od.contact_person ILIKE :q
                OR od.tel_number     ILIKE :q
                OR od.email          ILIKE :q
                OR od.mpan_mpr       ILIKE :q
            )
            ORDER BY od.created_at DESC
            LIMIT 50
        """), {'tenant_id': tenant_id, 'service_id': service_id, 'q': like}).mappings().all()

        return jsonify(_rows_to_list(rows)), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


@crm_bp.route('/leads/performance', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_performance():
    session = SessionLocal()
    try:
        tenant_id     = g.tenant_id
        current_user  = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1
        is_admin      = _is_admin_from_db(current_user)
        _raw_emp      = getattr(current_user, 'employee_id', None)
        employee_id   = int(_raw_emp) if _raw_emp is not None else None

        sql = """
            SELECT sm.stage_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master" sm ON od.stage_id = sm.stage_id
            WHERE od.tenant_id  = :tenant_id
            AND   od.service_id = :service_id
        """
        params = {'tenant_id': tenant_id, 'service_id': service_id}

        # ✅ FIX: Non-admin should only see own performance
        if employee_id:
            sql += ' AND od.opportunity_owner_employee_id = :employee_id'
            params['employee_id'] = employee_id

        rows = session.execute(text(sql), params).mappings().all()

        counts = dict(converted=0, renewed=0, renewed_directly=0, end_date_changed=0,
                      priced=0, in_progress=0, lost=0, not_contacted=0)

        for r in rows:
            stage = (r.get('stage_name') or '').lower()
            if stage == 'converted':                                              counts['converted'] += 1
            elif stage in ['already renewed', 'renewed']:                         counts['renewed'] += 1
            elif stage == 'renewed directly':                                     counts['renewed_directly'] += 1
            elif stage == 'end date changed':                                     counts['end_date_changed'] += 1
            elif stage == 'priced':                                               counts['priced'] += 1
            elif stage in ['callback','not answered','broker in place',
                           'email only','complaint','incorrect supplier']:         counts['in_progress'] += 1
            elif stage in ['lost','lost cot','invalid number','meter de-energised']: counts['lost'] += 1
            else:                                                                 counts['not_contacted'] += 1

        total = len(rows)
        success = round(
            (counts['converted'] + counts['renewed'] + counts['renewed_directly']) / total * 100, 1
        ) if total else 0

        return jsonify({
            'converted_count':        counts['converted'],
            'renewed_count':          counts['renewed'],
            'renewed_directly_count': counts['renewed_directly'],
            'end_date_changed_count': counts['end_date_changed'],
            'priced_count':           counts['priced'],
            'contacted_count':        counts['in_progress'],
            'not_contacted_count':    counts['not_contacted'],
            'lost_count':             counts['lost'],
            'success_rate':           success,
            'total_customers':        total,
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


@crm_bp.route('/leads/stats-by-employee', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_stats_by_employee():
    from backend.crm.supabase_client import get_supabase_client
    import logging
    logger = logging.getLogger(__name__)

    try:
        tenant_id    = g.tenant_id
        current_user = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1
        employee_id   = getattr(current_user, 'employee_id', None)

        logger.warning(
            '🔍 stats-by-employee: tenant_id=%s service_id=%s employee_id=%s',
            tenant_id, service_id, employee_id
        )

        if not employee_id:
            return jsonify({'stats': []}), 200

        db = get_supabase_client()

        # ✅ Everyone sees only their own count (mirrors renewals stats-by-employee)
        rows = db.execute_query('''
            SELECT
                em."employee_id",
                em."employee_name",
                COUNT(od."opportunity_id") AS count
            FROM "StreemLyne_MT"."Opportunity_Details" od
            JOIN "StreemLyne_MT"."Employee_Master" em
                ON od."opportunity_owner_employee_id" = em."employee_id"
            WHERE od."tenant_id" = %s
            AND od."service_id" = %s
            AND od."opportunity_owner_employee_id" = %s
            AND (od."is_allocated" = FALSE OR od."is_allocated" IS NULL)
            GROUP BY em."employee_id", em."employee_name"
        ''', (tenant_id, service_id, employee_id))

        stats = [
            {
                'employee_id':   r.get('employee_id'),
                'employee_name': r.get('employee_name'),
                'count':         int(r.get('count') or 0),
            }
            for r in (rows or [])
        ]

        logger.warning('📊 stats-by-employee result: %s', stats)
        return jsonify({'stats': stats}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'stats': []}), 500


@crm_bp.route('/leads/bulk-delete', methods=['POST'])
@require_tenant
def bulk_delete_leads():
    """Bulk delete multiple leads"""
    # ✅ Add logging to verify decorator worked
    import logging
    from flask import g
    
    logger = logging.getLogger(__name__)
    logger.info('Route handler: tenant_id=%s', g.get('tenant_id'))
    
    return crm_controller.bulk_delete_leads()


@crm_bp.route('/leads/table', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_table():
    return crm_controller.get_leads_table()


@crm_bp.route('/leads/import/preview', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads_preview():
    return crm_controller.import_leads_preview()


@crm_bp.route('/leads/import/confirm', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads_confirm():
    return crm_controller.import_leads_confirm()


@crm_bp.route('/leads/recycle-bin', methods=['GET'])
@token_required
@tenant_from_jwt
def get_recycle_bin():
    return crm_controller.get_recycle_bin()


@crm_bp.route('/leads/cleanup', methods=['PATCH'])
@token_required
@tenant_from_jwt
def delete_expired_lost_leads():
    return crm_controller.delete_expired_lost_leads()

def recalculate_lead_display_order(session, tenant_id, employee_id=None):
    """
    Recalculate display_order starting from 1 PER EMPLOYEE.
    Uses ROW_NUMBER() OVER (PARTITION BY opportunity_owner_employee_id ORDER BY created_at)
    so each salesperson's list always starts at 1.
    
    Args:
        session: SQLAlchemy session
        tenant_id: Tenant ID
        employee_id: Optional - recalculate only for this employee
    """
    if employee_id:
        # Recalculate only for this specific employee
        session.execute(text("""
            UPDATE "StreemLyne_MT"."Opportunity_Details" od
            SET display_order = sub.rn
            FROM (
                SELECT opportunity_id,
                       ROW_NUMBER() OVER (ORDER BY created_at ASC) AS rn
                FROM "StreemLyne_MT"."Opportunity_Details"
                WHERE tenant_id = :tenant_id
                  AND opportunity_owner_employee_id = :employee_id
                  AND (is_allocated = FALSE OR is_allocated IS NULL)
            ) sub
            WHERE od.opportunity_id = sub.opportunity_id
        """), {'tenant_id': tenant_id, 'employee_id': employee_id})
    else:
        # Recalculate for ALL employees at once using PARTITION BY
        session.execute(text("""
            UPDATE "StreemLyne_MT"."Opportunity_Details" od
            SET display_order = sub.rn
            FROM (
                SELECT opportunity_id,
                       ROW_NUMBER() OVER (
                           PARTITION BY opportunity_owner_employee_id
                           ORDER BY created_at ASC
                       ) AS rn
                FROM "StreemLyne_MT"."Opportunity_Details"
                WHERE tenant_id = :tenant_id
                  AND (is_allocated = FALSE OR is_allocated IS NULL)
                  AND opportunity_owner_employee_id IS NOT NULL
            ) sub
            WHERE od.opportunity_id = sub.opportunity_id
        """), {'tenant_id': tenant_id})
    
    session.flush()
    logging.getLogger(__name__).info(
        f"✅ Recalculated lead display_order per-employee "
        f"(tenant={tenant_id}, employee={employee_id or 'ALL'})"
    )

@crm_bp.route('/leads/import', methods=['POST'])
@token_required
@tenant_from_jwt
def import_leads():
    """
    ✅ FIXED: Ensures is_allocated = FALSE for imported leads + recalculates display_order
    """
    session = SessionLocal()
    try:
        tenant_id     = g.tenant_id
        service_param = request.args.get('service', 'electricity')
        service_id    = 2 if (service_param or '').strip().lower() == 'water' else 1
 
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided',
                            'total_rows': 0, 'successful': 0, 'failed': 1,
                            'errors': ['No file uploaded']}), 400
 
        file           = request.files.get('file')
        preview_result = crm_controller.crm_service.preview_lead_import(tenant_id, file)
 
        if not preview_result.get('success'):
            return jsonify({'success': False,
                            'message': preview_result.get('message', 'Validation failed'),
                            'total_rows': preview_result.get('total_rows', 0),
                            'successful': 0, 'failed': preview_result.get('total_rows', 1),
                            'errors': preview_result.get('errors', ['Validation failed'])}), 400
 
        if not preview_result.get('valid_rows', 0):
            return jsonify({'success': False, 'message': 'No valid rows to import',
                            'total_rows': preview_result.get('total_rows', 0),
                            'successful': 0, 'failed': preview_result.get('invalid_rows', 0),
                            'errors': preview_result.get('errors', ['No valid data found'])}), 400
 
        validated_data = [r['data'] for r in preview_result.get('rows', []) if r.get('is_valid')]
        created_by     = getattr(request.current_user, 'id', None)
        importing_employee_id = getattr(request.current_user, 'employee_id', None)
        
        confirm_result = crm_controller.crm_service.confirm_lead_import(
            tenant_id, validated_data, created_by, service_id)
 
        if 'success' in confirm_result and not confirm_result['success']:
            return jsonify({'success': False,
                            'message': confirm_result.get('message', 'Import failed'),
                            'total_rows': preview_result.get('total_rows', 0),
                            'successful': 0, 'failed': preview_result.get('total_rows', 0),
                            'errors': [confirm_result.get('error', 'Import failed')]}), 400
 
        inserted = confirm_result.get('inserted', 0)
        
        # ✅ CRITICAL: Force is_allocated = FALSE for newly imported leads
        if inserted > 0:
            try:
                # Update all recently created leads by this employee
                session.execute(text("""
                    UPDATE "StreemLyne_MT"."Opportunity_Details"
                    SET is_allocated = FALSE
                    WHERE tenant_id = :tenant_id
                      AND opportunity_owner_employee_id = :employee_id
                      AND created_at >= NOW() - INTERVAL '2 minutes'
                      AND (is_allocated IS NULL OR is_allocated = TRUE)
                """), {'tenant_id': tenant_id, 'employee_id': importing_employee_id})
                session.commit()
                logger.info(f'✅ Set is_allocated = FALSE for {inserted} imported leads')
                
                # Recalculate display_order
                recalculate_lead_display_order(session, tenant_id, importing_employee_id)
                session.commit()
                logger.info(f'✅ Recalculated display_order after importing {inserted} leads')
            except Exception as e:
                logger.error(f'❌ Post-import fix error: {e}')
                session.rollback()
        
        return jsonify({
            'success': inserted > 0,
            'message': f"Successfully imported {inserted} lead(s)" if inserted else "No new leads imported",
            'total_rows': preview_result.get('total_rows', 0),
            'successful': inserted,
            'failed': confirm_result.get('skipped', 0),
            'errors': confirm_result.get('errors', [])
        }), 200
 
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': str(e), 'total_rows': 0,
                        'successful': 0, 'failed': 1, 'errors': [str(e)]}), 500
    finally:
        session.close()


@crm_bp.route('/leads/import/template', methods=['GET'])
def download_leads_template():
    try:
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Import"
        
        # ✅ Headers match what preview_lead_import expects (case-insensitive)
        headers = [
            'Business Name',      # Matches 'business name'
            'Contact Person',     # Matches 'contact person'  
            'Tel Number',         # Matches 'tel number' ✓
            'Email',
            'MPAN_MPR',
            'Start Date',         # Matches 'start date' ✓
            'End Date',           # Matches 'end date' ✓
            'Annual Usage',
            'Address',
            'Site Address'
        ]
        ws.append(headers)
        
        # ✅ Example row with VALID data that will pass all validation checks
        ws.append([
            'Acme Corp',           # Business Name ✓
            'John Doe',            # Contact Person ✓
            '02071234567',         # Tel Number ✓ (UK format, 11 digits)
            'john@acme.com',       # Email
            '1234567890123',       # MPAN_MPR (13 digits)
            '01/01/2024',          # Start Date ✓ (DD/MM/YYYY UK format)
            '31/12/2024',          # End Date ✓ (DD/MM/YYYY UK format)
            '50000',               # Annual Usage
            '123 Main St, London', # Address
            '456 Business Park, London'  # Site Address
        ])
        
        # Style the header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Auto-adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, 
            download_name='leads_import_template.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@crm_bp.route('/leads/customer-type', methods=['GET'])
@token_required
@tenant_from_jwt
def get_leads_by_customer_type():
    return crm_controller.get_leads_by_customer_type()


@crm_bp.route('/leads/allocated', methods=['GET'])
@token_required
@tenant_from_jwt
def get_allocated_leads():
    session = SessionLocal()
    try:
        tenant_id     = g.tenant_id
        current_user  = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1
        _raw_emp      = getattr(current_user, 'employee_id', None)
        employee_id   = int(_raw_emp) if _raw_emp is not None else None

        if not employee_id:
            return jsonify([]), 200

        rows = session.execute(text("""
            SELECT od.*, sm.stage_name,
                   em.employee_name AS assigned_to_name,
                   COALESCE(od.business_name, od.opportunity_title) AS business_name,
                   sup.supplier_company_name AS supplier_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id   = sm.stage_id
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
            WHERE od.tenant_id  = :tenant_id
            AND   od.service_id = :service_id
            AND   od.opportunity_owner_employee_id = :employee_id
            AND   od.is_allocated = TRUE
            ORDER BY od.created_at DESC
        """), {'tenant_id': tenant_id, 'service_id': service_id, 'employee_id': employee_id}).mappings().all()

        return jsonify(_rows_to_list(rows)), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


@crm_bp.route('/leads/archives', methods=['GET'])
@token_required
@tenant_from_jwt
def get_archived_leads():
    return jsonify([]), 200


@crm_bp.route('/leads/<int:opportunity_id>/callback', methods=['POST', 'OPTIONS'])
@token_required
@tenant_from_jwt
def leads_callback(opportunity_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        data      = request.get_json(force=True, silent=True) or {}
        status    = data.get('status')

        if not status:
            return jsonify({'error': 'Status is required'}), 400

        STATUS_CFG = {
            "Callback":           {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "Not Answered":       {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "Priced":             {"deletes_record": False, "requires_notes": False, "requires_sold": True},
            "Lost":               {"deletes_record": True,  "requires_notes": True,  "requires_sold": False},
            "Lost COT":           {"deletes_record": True,  "requires_notes": True,  "requires_sold": False},
            "Already Renewed":    {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "Invalid Number":     {"deletes_record": True,  "requires_notes": False, "requires_sold": False},
            "Meter De-energised": {"deletes_record": True,  "requires_notes": False, "requires_sold": False},
            "Broker in Place":    {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "End Date Changed":   {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "Complaint":          {"deletes_record": False, "requires_notes": True,  "requires_sold": False},
            "Email Only":         {"deletes_record": False, "requires_notes": False, "requires_sold": False},
            "Renewed Directly":   {"deletes_record": False, "requires_notes": True,  "requires_sold": False},
            "Incorrect Supplier": {"deletes_record": False, "requires_notes": True,  "requires_sold": False},
            "Converted":          {"deletes_record": False, "requires_notes": False, "requires_sold": False},
        }

        if status not in STATUS_CFG:
            return jsonify({'error': f'Invalid status: {status}'}), 400

        cfg      = STATUS_CFG[status]
        notes    = data.get('notes', '')
        is_sold  = data.get('is_sold')
        stage_id = data.get('stage_id')

        if cfg['requires_notes'] and not (notes or '').strip():
            return jsonify({'error': 'Notes are required for this status'}), 400
        if cfg['requires_sold'] and is_sold is None:
            return jsonify({'error': 'Please select if the contract was sold'}), 400
        if status == 'Already Renewed' and not data.get('renewed_by'):
            return jsonify({'error': 'Please select if renewed by customer or agent'}), 400

        lead = session.execute(text("""
            SELECT opportunity_id FROM "StreemLyne_MT"."Opportunity_Details"
            WHERE tenant_id = :t
            AND (tenant_lead_id = :id OR opportunity_id = :id)
            LIMIT 1
        """), {'t': tenant_id, 'id': opportunity_id}).mappings().first()

        if not lead:
            return jsonify({'error': 'Lead not found'}), 404

        real_id = lead['opportunity_id']

        if cfg['deletes_record']:
            if not stage_id:
                s = session.execute(text(
                    "SELECT stage_id FROM \"StreemLyne_MT\".\"Stage_Master\" "
                    "WHERE LOWER(stage_name) = :n LIMIT 1"
                ), {'n': status.lower()}).mappings().first()
                stage_id = s['stage_id'] if s else 5
            session.execute(text(
                'UPDATE "StreemLyne_MT"."Opportunity_Details" SET stage_id = :s '
                'WHERE opportunity_id = :id AND tenant_id = :t'
            ), {'s': stage_id, 'id': real_id, 't': tenant_id})
            session.commit()
            return jsonify({'success': True, 'moved_to_recycle_bin': True,
                            'message': f'Moved to recycle bin ({status})'}), 200

        if status == 'Priced' and is_sold is False:
            session.execute(text(
                'UPDATE "StreemLyne_MT"."Opportunity_Details" SET stage_id = :s '
                'WHERE opportunity_id = :id AND tenant_id = :t'
            ), {'s': stage_id or 4, 'id': real_id, 't': tenant_id})
            session.commit()
            return jsonify({'success': True, 'moved_to_priced': True}), 200

        new_end = data.get('new_end_date')
        if new_end and status in ('End Date Changed', 'Already Renewed'):
            session.execute(text(
                'UPDATE "StreemLyne_MT"."Opportunity_Details" SET end_date = :d '
                'WHERE opportunity_id = :id AND tenant_id = :t'
            ), {'d': new_end, 'id': real_id, 't': tenant_id})

        new_supplier = (data.get('new_supplier') or '').strip()
        if new_supplier:
            sup = session.execute(text(
                'SELECT supplier_id FROM "StreemLyne_MT"."Supplier_Master" '
                'WHERE LOWER(supplier_company_name) = LOWER(:n) LIMIT 1'
            ), {'n': new_supplier}).mappings().first()
            if sup:
                session.execute(text(
                    'UPDATE "StreemLyne_MT"."Opportunity_Details" SET supplier_id = :s '
                    'WHERE opportunity_id = :id AND tenant_id = :t'
                ), {'s': sup['supplier_id'], 'id': real_id, 't': tenant_id})

        if stage_id:
            session.execute(text(
                'UPDATE "StreemLyne_MT"."Opportunity_Details" SET stage_id = :s '
                'WHERE opportunity_id = :id AND tenant_id = :t'
            ), {'s': stage_id, 'id': real_id, 't': tenant_id})

        session.commit()
        return jsonify({'success': True, 'message': 'Callback saved successfully', 'status': status}), 200

    except Exception as e:
        session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


@crm_bp.route('/leads/priced', methods=['GET'])
@token_required
@tenant_from_jwt
def get_priced_leads():
    session = SessionLocal()
    try:
        tenant_id     = g.tenant_id
        current_user  = request.current_user
        service_param = request.args.get('service', 'utilities')
        service_id    = 2 if service_param.strip().lower() == 'water' else 1
        is_admin      = _is_admin_from_db(current_user)
        _raw_emp      = getattr(current_user, 'employee_id', None)
        employee_id   = int(_raw_emp) if _raw_emp is not None else None
        salesperson   = request.args.get('salesperson')

        sql = """
            SELECT od.*, sm.stage_name,
                   em.employee_name AS assigned_to_name,
                   COALESCE(od.business_name, od.opportunity_title) AS business_name,
                   sup.supplier_company_name AS supplier_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id   = sm.stage_id
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
            WHERE od.tenant_id  = :tenant_id
            AND   od.service_id = :service_id
            AND   LOWER(sm.stage_name) = 'priced'
        """
        params = {'tenant_id': tenant_id, 'service_id': service_id}

        if is_admin and salesperson and salesperson != 'All':
            try:
                sql += ' AND od.opportunity_owner_employee_id = :salesperson'
                params['salesperson'] = int(salesperson)
            except ValueError:
                pass
        elif not is_admin and employee_id:
            sql += ' AND od.opportunity_owner_employee_id = :employee_id'
            params['employee_id'] = employee_id

        sql += ' ORDER BY od.created_at DESC'
        rows = session.execute(text(sql), params).mappings().all()
        return jsonify(_rows_to_list(rows)), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


# ========================================
# CLIENT ROUTES
# ========================================

@crm_bp.route('/clients', methods=['POST'])
@require_tenant
def create_client():
    return crm_controller.create_client()


@crm_bp.route('/clients/<int:client_id>/call-summary', methods=['POST'])
@require_tenant
def create_call_summary(client_id):
    return crm_controller.create_call_summary(client_id)


@crm_bp.route('/clients/<int:client_id>/upload', methods=['POST'])
@require_tenant
def client_upload_document(client_id):
    return crm_controller.client_upload_document(client_id)


# ========================================
# PROJECT ROUTES
# ========================================

@crm_bp.route('/projects', methods=['GET'])
@require_tenant
def get_projects():
    return crm_controller.get_projects()


@crm_bp.route('/projects/<int:project_id>', methods=['GET'])
@require_tenant
def get_project_detail(project_id):
    return crm_controller.get_project_detail(project_id)


# ========================================
# DEAL ROUTES
# ========================================

@crm_bp.route('/deals', methods=['GET'])
@require_tenant
def get_deals():
    return crm_controller.get_deals()


@crm_bp.route('/deals/<int:contract_id>', methods=['GET'])
@require_tenant
def get_deal_detail(contract_id):
    return crm_controller.get_deal_detail(contract_id)


# ========================================
# USER ROUTES
# ========================================

@crm_bp.route('/users', methods=['GET'])
@require_tenant
def get_users():
    return crm_controller.get_users()


@crm_bp.route('/employees', methods=['GET'])
@token_required
@tenant_from_jwt
def get_employees():
    return crm_controller.get_employees()


# ========================================
# SUPPORTING DATA
# ========================================

@crm_bp.route('/roles', methods=['GET'])
def get_roles():
    return crm_controller.get_roles()


@crm_bp.route('/stages', methods=['GET'], strict_slashes=False)
@token_required
@tenant_from_jwt
def get_stages():
    return crm_controller.get_stages()


@crm_bp.route('/services', methods=['GET'])
def get_services():
    return crm_controller.get_services()


@crm_bp.route('/suppliers', methods=['GET'])
@require_tenant
def get_suppliers():
    return crm_controller.get_suppliers()


@crm_bp.route('/interactions', methods=['GET'])
@require_tenant
def get_interactions():
    return crm_controller.get_interactions()


# ========================================
# DASHBOARD
# ========================================

@crm_bp.route('/dashboard', methods=['GET'])
@require_tenant
def get_dashboard():
    return crm_controller.get_dashboard()


@crm_bp.route('/priced', methods=['GET'])
@token_required
@tenant_from_jwt
def get_priced():
    return crm_controller.get_priced()


@crm_bp.route('/cleansing', methods=['GET'])
@token_required
@tenant_from_jwt
def get_cleansing():
    session = SessionLocal()
    try:
        tenant_id = g.tenant_id
        records   = []

        lead_rows = session.execute(text("""
            SELECT
                od.opportunity_id AS id, od.opportunity_id AS client_id,
                od.tenant_lead_id AS display_id, od.tenant_lead_id AS display_order,
                COALESCE(od.business_name, od.opportunity_title) AS business_name,
                od.contact_person, od.tel_number AS phone, od.mobile_no,
                od.mpan_mpr, od.mpan_mpr AS mpan_top, od.supplier_id,
                sup.supplier_company_name AS supplier_name, od.annual_usage,
                od.start_date, od.end_date,
                sm.stage_name AS cleansing_reason, od.created_at AS flagged_at,
                od.notes, od.opportunity_owner_employee_id AS assigned_to_id,
                em.employee_name AS assigned_to_name
            FROM "StreemLyne_MT"."Opportunity_Details" od
            LEFT JOIN "StreemLyne_MT"."Stage_Master"    sm  ON od.stage_id    = sm.stage_id
            LEFT JOIN "StreemLyne_MT"."Supplier_Master" sup ON od.supplier_id = sup.supplier_id
            LEFT JOIN "StreemLyne_MT"."Employee_Master" em  ON od.opportunity_owner_employee_id = em.employee_id
            WHERE od.tenant_id = :t
            AND sm.stage_name IN ('Invalid Number', 'Incorrect Supplier')
            ORDER BY od.created_at DESC
        """), {'t': tenant_id}).mappings().all()

        for r in lead_rows:
            rd = dict(r)
            records.append({k: _serial(v) for k, v in rd.items()} | {'source': 'lead'})

        try:
            from backend.models import Client_Master, Energy_Contract_Master, Project_Details, Supplier_Master
            client_rows = (
                session.query(Client_Master, Energy_Contract_Master, Supplier_Master)
                .outerjoin(Project_Details, Client_Master.client_id == Project_Details.client_id)
                .outerjoin(Energy_Contract_Master, Project_Details.project_id == Energy_Contract_Master.project_id)
                .outerjoin(Supplier_Master, Energy_Contract_Master.supplier_id == Supplier_Master.supplier_id)
                .filter(
                    Client_Master.tenant_id == tenant_id,
                    Client_Master.is_deleted == True,
                    Client_Master.deleted_reason.in_(['Invalid Number', 'Incorrect Supplier']),
                ).all()
            )
            for client, contract, supplier in client_rows:
                records.append({
                    'id': client.client_id, 'client_id': client.client_id,
                    'display_id': getattr(client, 'display_id', None),
                    'display_order': getattr(client, 'display_order', None),
                    'business_name': getattr(client, 'client_company_name', None) or 'Unknown',
                    'contact_person': getattr(client, 'client_contact_name', None),
                    'phone': getattr(client, 'client_phone', None),
                    'mobile_no': getattr(client, 'mobile_no', None),
                    'mpan_mpr': getattr(contract, 'mpan_number', None) if contract else None,
                    'mpan_top': getattr(contract, 'mpan_number', None) if contract else None,
                    'supplier_id': contract.supplier_id if contract else None,
                    'supplier_name': supplier.supplier_company_name if supplier else None,
                    'annual_usage': None,
                    'start_date': contract.contract_start_date.isoformat() if contract and contract.contract_start_date else None,
                    'end_date': contract.contract_end_date.isoformat() if contract and contract.contract_end_date else None,
                    'cleansing_reason': client.deleted_reason,
                    'flagged_at': client.deleted_at.isoformat() if client.deleted_at else None,
                    'notes': getattr(client, 'deleted_notes', None),
                    'assigned_to_id': getattr(client, 'assigned_to_id', None),
                    'assigned_to_name': None, 'source': 'energy_client',
                })
        except Exception as ec_err:
            logger.warning('Could not load energy clients for cleansing: %s', ec_err)

        records.sort(key=lambda x: x.get('flagged_at') or '', reverse=True)
        return jsonify({'records': records, 'total': len(records)}), 200

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        try: session.close()
        except Exception: pass


# ========================================
# HEALTH CHECK
# ========================================

@crm_bp.route('/health', methods=['GET'])
def health_check():
    return {'success': True, 'module': 'CRM', 'status': 'operational'}, 200


@crm_bp.route('/debug/tenant/<int:tenant_id>', methods=['GET'])
def debug_tenant_lookup(tenant_id):
    try:
        from backend.crm.repositories.tenant_repository import TenantRepository
        repo = TenantRepository()
        tenant = repo.get_tenant_by_id(tenant_id)
        return {'success': bool(tenant), 'tenant_id_requested': tenant_id,
                'tenant_found': tenant is not None, 'tenant_data': tenant}, 200 if tenant else 404
    except Exception as e:
        import traceback
        return {'success': False, 'error': str(e), 'traceback': traceback.format_exc()}, 500